from django.db import models
import logging
from django.contrib.auth.models import User
import os
from seq_common.utils import classes
from openpyxl.reader.excel import load_workbook
from django.template import loader
from django.template.context import Context

MAIN_PATH = 'c:\\DEV\\Sources\\gso_sellmystuff\\resources'
STATICS_PATH = 'c:\\DEV\\Sources\\gso_sellmystuff\\templates\\statics'

LOGGER = logging.getLogger(__name__)

def setup():
    populate_attributes_from_xlsx('sellmystuff.models.Attributes', os.path.join(MAIN_PATH,'Repository Setup.xlsx'))
    generate_attributes()
    # populate_attributes_from_xlsx('universe.models.Dictionary', os.path.join(MAIN_PATH,'Repository Setup.xlsx'))

def generate_attributes():
    all_types = Attributes.objects.all().order_by('type').distinct('type')
    for a_type in all_types:
        all_elements = Attributes.objects.filter(type=a_type.type, active=True)
        context = Context({"selection": all_elements})
        template = loader.get_template('helper/attributes_option_renderer.html')
        rendition = template.render(context)
        outfile = os.path.join(STATICS_PATH, a_type.type + '_en.html')
        with open(outfile,'w') as o:
            o.write(rendition.encode('utf-8'))

def populate_attributes_from_xlsx(model_name, xlsx_file):
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 0
    # Reading header
    header = []
    for column_index in range(0, sheet.get_highest_column()):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<sheet.get_highest_row():
        if model.objects.filter(identifier=sheet.cell(row = row_index, column=0).value).exists():
            instance = model.objects.get(identifier=sheet.cell(row = row_index, column=0).value)
        else:
            instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column=i).value
            setattr(instance, header[i], value)
        instance.save()
        row_index += 1

def populate_model_from_xlsx(model_name, xlsx_file):
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 0
    # Reading header
    header = []
    for column_index in range(0, sheet.get_highest_column()):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<sheet.get_highest_row():
        instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column=i).value
            field_info = Attributes()
            field_info.short_name = header[i]
            field_info.name = header[i]
            instance.set_attribute('excel', field_info, value)
        instance.save()
        row_index += 1
class CoreModel(models.Model):

    def get_editable_fields(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()!='ManyToManyField':
                values.append(field)
        return values
        
    def get_associable_field(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()=='ManyToManyField':
                values.append(field)
        return values        

    def get_fields(self):
        return []
    
    def get_identifier(self):
        return 'name'
    
    def list_values(self):
        values = []
        for field in self.get_fields():
            LOGGER.debug(self.__class__.__name__ + ' * ' + field)
            if field in self._meta.get_all_field_names():
                if self._meta.get_field(field).get_internal_type()=='ManyToManyField' and getattr(self,field)!=None:
                    values.append(str([e.list_values() for e in list(getattr(self,field).all())]))
                elif self._meta.get_field(field).get_internal_type()=='ForeignKey' and getattr(self,field)!=None:
                    values.append(getattr(self,field).get_value())
                else:
                    values.append(str(getattr(self,field)))
            else:
                # Generic foreign key
                values.append(getattr(self,field).get_value())
        return values
    
    def get_value(self):
        if self.get_identifier()!=None:
            return getattr(self, self.get_identifier())
        else:
            return None
        
    def __unicode__(self):
        return unicode(self.get_value())
    
    class Meta:
        ordering = ['id']
    
class Attributes(CoreModel):
    identifier = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=32)
    type = models.CharField(max_length=64)
    active = models.BooleanField()
    
    def get_fields(self):
        return ['identifier','name','short_name','type','active']
    
    class Meta:
        ordering = ['name']


class Account(CoreModel):
    owner = models.ForeignKey(User, related_name='account_owner_rel')
    title = models.TextField(max_length=128)
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'account_type'}, related_name='account_type_rel')
    creation_date = models.DateTimeField()
    last_update = models.DateTimeField()
    expiry_date = models.DateTimeField(null=True, blank=True)
    currency = models.ForeignKey(Attributes, limit_choices_to={'type':'currency'}, related_name='account_currency_rel')
    country = models.ForeignKey(Attributes, limit_choices_to={'type':'country_iso2'}, related_name='account_country_rel')
    active = models.BooleanField()

    def get_fields(self):
        return ['owner','title','type','creation_date','last_update','expiry_date','currency','active']
    
    class Meta:
        ordering = ['title']

class Comment(CoreModel):
    content = models.TextField()
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'comment_type'}, related_name='comment_type_rel')
    reference = models.ForeignKey("Comment", related_name='reference_comment_rel', null=True)
    publish_date = models.DateTimeField()
    last_update = models.DateTimeField()
    active = models.BooleanField()
    
    owner = models.ForeignKey(Account, related_name='comment_owner_rel')

class Media(CoreModel):
    title = models.CharField(max_length=256)
    description = models.TextField(null=True, blank=True)
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'media_type'}, related_name='media_type_rel')
    identifier = models.CharField(max_length=512)
    last_update = models.DateTimeField()
    active = models.BooleanField()
        
    owner = models.ForeignKey(Account, related_name='media_owner_rel')
    
    def get_fields(self):
        return ['title','description','type','identifier','last_update']
    
    class Meta:
        ordering = ['title']

class Advertisement(CoreModel):
    title = models.CharField(max_length=256)
    sub_title = models.CharField(max_length=128, null=True, blank=True)
    publish_date = models.DateTimeField(null=True, blank=True)
    last_update = models.DateTimeField()
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'ad_type'}, related_name='ad_type_rel')
    category = models.ForeignKey(Attributes, limit_choices_to={'type':'ad_cat_type'}, related_name='ad_cat_type_rel')
    medias = models.ManyToManyField("Media", related_name='ad_media_rel', null=True)
    comments = models.ManyToManyField("Comment", related_name='ad_comment_rel', null=True)
    price = models.FloatField(null=True, blank=True)
    views = models.IntegerField(default=0)
    
    sold = models.BooleanField()
    buyer = models.ForeignKey(Account, related_name='ad_buyer_rel', null=True)
    active = models.BooleanField()
    owner = models.ForeignKey(Account, related_name='ad_owner_rel')
    
    def get_fields(self):
        return ['title','sub_title','publish_date','last_update','type','category','medias','comments','price','sold','buyer','active']
    
    class Meta:
        ordering = ['title']
        
class Offer(CoreModel):
    owner = models.ForeignKey(Account, related_name='offer_owner_rel')
    advertisement = models.ForeignKey(Advertisement, related_name='offer_ad_rel')
    publish_date = models.DateTimeField()
    last_update = models.DateTimeField()
    price = models.FloatField(null=True, blank=True)
    
    def get_fields(self):
        return ['owner','advertisement','publish_date','last_update','price']
    
    class Meta:
        ordering = ['publish_date']